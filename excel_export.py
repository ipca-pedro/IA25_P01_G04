"""
Exportador Excel para horários CSP
Gera planilha semanal com cores por turma
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from csp_formulation import get_day

def export_to_excel(solution, dataset, filename="horario_semanal.xlsx"):
    """Exporta horário com uma tabela por turma, cada uma com sua cor"""
    
    # Cores por turma
    colors = ["FFE6E6", "E6F3FF", "E6FFE6", "FFFFE6", "F0E6FF", "FFE6F0"]
    
    days = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    slots = ["9h-11h", "11h-13h", "14h-16h", "16h-18h"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Horários por Turma"
    
    current_row = 1
    
    for class_idx, class_name in enumerate(dataset['classes']):
        # Título da turma
        ws.cell(current_row, 1, f"TURMA {class_name}").font = Font(bold=True, size=14)
        current_row += 1
        
        # Cabeçalho da tabela
        headers = ["Horário"] + days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
        
        # Cria grid da turma
        class_schedule = {}
        for day in range(5):
            for slot in range(4):
                class_schedule[(day, slot)] = ""
        
        # Preenche com aulas da turma
        for course in dataset['cc'][class_name]:
            for lesson in [1, 2]:
                if (course, lesson) in solution:
                    slot, room = solution[(course, lesson)]
                    day_idx = get_day(slot) - 1
                    slot_idx = (slot - 1) % 4
                    
                    # Encontra professor da UC
                    teacher = None
                    for prof, courses in dataset['dsd'].items():
                        if course in courses:
                            teacher = prof
                            break
                    
                    class_schedule[(day_idx, slot_idx)] = f"{course}_L{lesson}\n{teacher}\n[{room}]"
        
        # Preenche tabela da turma
        color = colors[class_idx % len(colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        for slot_idx, slot_time in enumerate(slots):
            # Coluna horário
            ws.cell(current_row, 1, slot_time)
            
            # Colunas dos dias
            for day_idx in range(5):
                cell = ws.cell(current_row, day_idx + 2, class_schedule[(day_idx, slot_idx)])
                if class_schedule[(day_idx, slot_idx)]:  # Se tem aula
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            current_row += 1
        
        # Espaço entre turmas
        current_row += 2
    
    # Ajusta larguras
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(filename)
    print(f"[OK] Horário por turmas exportado para: {filename}")
    return filename